"""
Collect student details and marks from all the *.xlsx files in a directory
Using openpyxl but it can only read the value of very simple calcuated cell
Even =SUM(...) does not work :'(
Ref: https://stackoverflow.com/questions/23350581/openpyxl-1-8-5-reading-the-result-of-a-formula-typed-in-a-cell-using-openpyxl

Please use markscollector.py instead that employs the xlwinds instead
"""

from openpyxl import load_workbook
import pandas as pd
import os

df = pd.DataFrame(columns=['StudID','StudName','Class','Mark'])
i = 0

directory = os.fsencode(".")
    
for file in os.listdir(directory):
	filename = os.fsdecode(file)
	if filename.endswith(".xlsx"):
		print (filename)

		# Read xlsx
		wb = load_workbook(filename = filename, data_only=True)
		ws = wb.active
		#print (ws['B1'].value)
		
		row = []
		row.append(ws['B1'].value)
		row.append(ws['B2'].value)
		row.append(ws['B3'].value)
		row.append(ws['B10'].value)
		df.loc[i] = row
		i += 1
	
print(df)
		
		
      
