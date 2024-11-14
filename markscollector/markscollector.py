"""
Collect student details and marks from all the *.xlsx files in a directory
Using openpyxl 
"""

from openpyxl import load_workbook
import pandas as pd
import os

sheet_cond = "Ass1.xlsx"
mark_filename = "StudMark.xlsx"

directory = os.fsencode(".")

df = pd.DataFrame(columns=['StudID','StudName','Class','Mark'])
i = 0    
for file in os.listdir(directory):
	filename = os.fsdecode(file)
	if filename.endswith(sheet_cond):
		print ('Collecting Total: ' + filename)

		# Read xlsx
		wb = load_workbook(filename = filename, data_only=True)
		ws = wb.active
		#print (ws['B1'].value)
		
		row = []
		row.append(ws['B2'].value)
		row.append(ws['B1'].value)
		row.append(ws['D2'].value)
		row.append(ws['D9'].value)
		df.loc[i] = row
		i += 1
		wb.close()
	
#print(df)
#df.to_csv(mark_filename)
df.to_excel(mark_filename)
		
		
      
