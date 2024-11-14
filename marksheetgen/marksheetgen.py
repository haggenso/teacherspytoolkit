"""
This program read in class roster from the file "classroster.csv" with the format:

StudID	StudName	Class
A001	Name1	A

The Excel template is in the file "marksheet_template.xlsx".
You can also switch between sheets with the statement
ws = wb[sheets[0]]

Ref: https://stackoverflow.com/questions/45756731/how-to-switch-between-sheets-in-excel-openpyxl-python-to-make-changes?rq=1

"""

from openpyxl import load_workbook
import pandas as pd

# Customise the filenames
# For Output Excel filenames, it requires some programming so not included here
in_filename = 'classroster.csv'
template_name ='marksheet_template.xlsx'

# Read Student Roster
df = pd.read_csv(in_filename)
# print(df)

# Loop
for index, row in df.iterrows():
	# print (row['StudID'],row['StudName'],row['Class'])

	# Read xlsx
	wb = load_workbook(filename = template_name, data_only=False)
	ws = wb.active
	# print(ws)

	ws['B1']=row['StudName']
	ws['B2']=row['StudID']
	ws['D2']=row['Class']
	wb.save(str(row['StudID']) + "_" + row['StudName'] + "_" + "Ass1.xlsx")
